from pathlib import Path
from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, file_path):
        self.file_path = Path(file_path)

    def read_summary(self):
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        rows = []

        for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True):
            rows.append(list(row))

        summary = {
            "file_name": self.file_path.name,
            "university_name": self.file_path.stem,
            "sheet_name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "rows": rows
        }

        wb.close()
        return summary

    def read_all_rows(self):
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

        wb.close()
        return rows