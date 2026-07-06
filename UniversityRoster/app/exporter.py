from pathlib import Path
from shutil import copyfile
from datetime import datetime

from openpyxl import Workbook, load_workbook

from app.config import APP_NAME, APP_VERSION, LOG_SHEET_NAME


class ExcelExporter:
    OUTPUT_COLUMNS = [
        "背番号", "氏名", "ふりがな", "略名", "スタッフ分類", "学年",
        "生年月日", "性別", "出身地", "出身小中学校", "高校", "大学",
        "身長cm", "体重kg", "守備位置", "投", "打", "経歴",
    ]

    def __init__(self, output_file, overwrite=True, template_file=None):
        self.output_file = Path(output_file)
        self.overwrite = overwrite
        self.template_file = Path(template_file) if template_file else None

    def prepare_output_file(self):
        if self.template_file and self.template_file.exists():
            if self.output_file.resolve() != self.template_file.resolve():
                copyfile(self.template_file, self.output_file)
            return load_workbook(self.output_file)

        if self.output_file.exists():
            return load_workbook(self.output_file)

        wb = Workbook()
        ws = wb.active
        ws.title = LOG_SHEET_NAME
        return wb

    def safe_sheet_name(self, name):
        invalid_chars = ["\\", "/", "?", "*", "[", "]", ":"]
        sheet_name = str(name)

        for ch in invalid_chars:
            sheet_name = sheet_name.replace(ch, "")

        return sheet_name[:31]

    def prepare_sheet(self, wb, university_name):
        sheet_name = self.safe_sheet_name(university_name)

        if sheet_name in wb.sheetnames:
            if self.overwrite:
                del wb[sheet_name]
            else:
                raise ValueError(f"既存シートがあります: {sheet_name}")

        ws = wb.create_sheet(sheet_name)
        return ws

    def make_output_record(self, record, university_name):
        staff_type = record.get("スタッフ分類", "")
        position = record.get("守備位置", "")

        if staff_type != "選手":
            position = ""

        return {
            "背番号": record.get("背番号", ""),
            "氏名": record.get("氏名", ""),
            "ふりがな": record.get("ふりがな", ""),
            "略名": record.get("略名", ""),
            "スタッフ分類": staff_type,
            "学年": record.get("学年", ""),
            "生年月日": record.get("生年月日", ""),
            "性別": record.get("性別", ""),
            "出身地": record.get("出身地", ""),
            "出身小中学校": record.get("出身小中学校", ""),
            "高校": record.get("高校", ""),
            "大学": university_name,
            "身長cm": record.get("身長cm", ""),
            "体重kg": record.get("体重kg", ""),
            "守備位置": position,
            "投": record.get("投", ""),
            "打": record.get("打", ""),
            "経歴": record.get("経歴", ""),
        }

    def write_university_sheet(self, wb, university_name, records):
        ws = self.prepare_sheet(wb, university_name)

        for col_index, header in enumerate(self.OUTPUT_COLUMNS, start=1):
            ws.cell(row=1, column=col_index, value=header)

        for row_index, record in enumerate(records, start=2):
            output_record = self.make_output_record(record, university_name)

            for col_index, header in enumerate(self.OUTPUT_COLUMNS, start=1):
                ws.cell(
                    row=row_index,
                    column=col_index,
                    value=output_record.get(header, "")
                )

        return ws

    def write_log_sheet(self, wb, university_records):
        if LOG_SHEET_NAME in wb.sheetnames:
            ws = wb[LOG_SHEET_NAME]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(LOG_SHEET_NAME, 0)

        total_universities = len(university_records)
        total_records = sum(len(records) for records in university_records.values())

        rows = [
            [APP_NAME],
            ["Version", APP_VERSION],
            ["実行日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["出力ファイル", str(self.output_file)],
            ["テンプレート", str(self.template_file) if self.template_file else ""],
            ["処理大学数", total_universities],
            ["総データ件数", total_records],
            [],
            ["大学名", "件数", "結果"],
        ]

        for university_name, records in university_records.items():
            rows.append([university_name, len(records), "OK"])

        for r, row_values in enumerate(rows, start=1):
            for c, value in enumerate(row_values, start=1):
                ws.cell(row=r, column=c, value=value)

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 16

    def save(self, university_records):
        wb = self.prepare_output_file()

        for university_name, records in university_records.items():
            self.write_university_sheet(wb, university_name, records)

        self.write_log_sheet(wb, university_records)

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        try:
            wb.save(self.output_file)
        except PermissionError:
            raise PermissionError(
                "出力ファイルがExcelで開かれている可能性があります。"
                "Excelを閉じてから再実行してください。"
            )