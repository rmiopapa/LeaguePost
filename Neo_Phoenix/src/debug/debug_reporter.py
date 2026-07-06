from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as e:
    raise ModuleNotFoundError("openpyxl がありません。python -m pip install openpyxl を実行してください。") from e


class DebugReporter:
    """
    Phoenix V2.6 Debug Edition Sprint4.3 ReasonPriorityRollback

    判定ロジックは触らず、Actual / Virtual の状態を1プレー単位で可視化する。
    """

    def write(self, day: Any, out_path: Path | str) -> Path:
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_debug = wb.create_sheet("DebugTrace")
        ws_run_detail = wb.create_sheet("RunDetail")
        ws_vdiff = wb.create_sheet("VirtualDiff")
        ws_team_pitcher = wb.create_sheet("TeamPitcherSummary")
        ws_runtime = wb.create_sheet("PitcherRuntimeDebug")
        ws_trace = wb.create_sheet("PlayTrace")
        ws_first = wb.create_sheet("FirstDifference")
        ws_scores = wb.create_sheet("ScoreJudgments")
        ws_warn = wb.create_sheet("Warnings")

        self._setup_styles()
        self._write_summary(ws_summary, day)
        self._write_debug_trace(ws_debug, day)
        run_rows = self._write_run_detail(ws_run_detail, day)
        self._write_virtual_diff(ws_vdiff, run_rows)
        self._write_team_pitcher_summary(ws_team_pitcher, run_rows)
        self._write_pitcher_runtime_debug(ws_runtime, day)
        first_rows = self._write_play_trace(ws_trace, day)
        self._write_first_difference(ws_first, first_rows)
        self._write_score_judgments(ws_scores, day)
        self._write_warnings(ws_warn, day)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            self._autosize(ws)

        wb.save(out_path)
        return out_path

    def _setup_styles(self):
        self.header_fill = PatternFill("solid", fgColor="1F4E78")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.ok_fill = PatternFill("solid", fgColor="E2F0D9")
        self.ng_fill = PatternFill("solid", fgColor="FCE4D6")
        self.warn_fill = PatternFill("solid", fgColor="FFF2CC")
        self.first_fill = PatternFill("solid", fgColor="FFD966")
        self.thin = Side(style="thin", color="D9E2F3")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)

    def _header(self, ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.border


    def _write_debug_trace(self, ws, day):
        from src.pitcher.pitcher_state import report_state_simple_text

        pitcher_columns = self._collect_pitcher_columns(day)
        self._header(ws, ["Inning", "Seq", "Event", "Actual", "Team Virtual"] + pitcher_columns + ["Δ", "得点", "Team判定", "Team理由", "Pitcher判定", "Pitcher責任/理由"])
        section_fill = PatternFill("solid", fgColor="E7E6E6")
        total_cols = 5 + len(pitcher_columns) + 6

        for game in getattr(day, "games", []):
            for half in getattr(game.analysis, "halves", []):
                title = str(getattr(half, "title", "") or "").replace("Actual ", "").replace("Virtual ", "")
                display_map = self._pitcher_display_map_for_half(half)
                ws.append([f"── {title} ──"] + [""] * (total_cols - 1))
                r = ws.max_row
                for c in range(1, total_cols + 1):
                    ws.cell(r, c).fill = section_fill
                    ws.cell(r, c).font = Font(bold=True)
                    ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(r, c).border = self.border

                display_map = self._pitcher_display_map_for_half(half)
                pitcher_by_runner = self._runner_pitcher_map(getattr(half.actual_report, "runner_history", []), display_map)
                compare = getattr(half, "compare_result", None)
                team_score_judgments = list(getattr(compare, "team_judgments", []) or getattr(compare, "judgments", []) or [])
                pitcher_score_judgments = list(getattr(compare, "pitcher_judgments", []) or team_score_judgments)
                score_cursor = 0

                actual_by_seq = {getattr(p, "seq", None): p for p in getattr(half.actual_report, "plays", [])}
                virtual_by_seq = {getattr(p, "seq", None): p for p in getattr(half.virtual_report, "plays", [])}
                pitcher_virtuals_by_seq = {}
                current_pitcher_by_seq = {}
                pitcher_virtual_report = getattr(half, "pitcher_virtual_report", None) or getattr(half, "virtual_report", None)
                for runtime_row in getattr(pitcher_virtual_report, "pitcher_runtime_debug", []):
                    parts = runtime_row.split(",", 8)
                    if len(parts) == 9:
                        try:
                            seq_key = int(parts[0])
                        except Exception:
                            continue
                        current_pitcher_by_seq[seq_key] = self._display_pitcher(parts[2], display_map)
                        pitcher_virtuals_by_seq[seq_key] = self._parse_all_pitcher_virtuals(parts[7], display_map)
                pitcher_changes = {}
                for row in getattr(half.actual_report, "pitcher_changes", []):
                    parts = row.split(",", 3)
                    if len(parts) == 4:
                        try:
                            pitcher_changes[int(parts[0])] = parts
                        except Exception:
                            pass
                seqs = sorted({s for s in set(actual_by_seq) | set(virtual_by_seq) | set(pitcher_changes) | set(pitcher_virtuals_by_seq) if s is not None})
                last_actual = ""
                last_virtual = ""
                last_pitcher_virtuals = {name: "" for name in pitcher_columns}

                for seq in seqs:
                    ap = actual_by_seq.get(seq)
                    vp = virtual_by_seq.get(seq)
                    pv_map = dict(last_pitcher_virtuals)
                    pv_map.update(pitcher_virtuals_by_seq.get(seq, {}))

                    if seq in pitcher_changes and not ap and not vp:
                        row = [title, seq, "投手交代", last_actual, last_virtual] + [pv_map.get(name, "") for name in pitcher_columns] + ["責任投手更新", "", "", "", "", ""]
                    else:
                        src = ap or vp
                        actual = report_state_simple_text(getattr(ap, "after_text", ""), getattr(ap, "outs_after", 0)) if ap else last_actual
                        virtual = report_state_simple_text(getattr(vp, "after_text", ""), getattr(vp, "outs_after", 0)) if vp else last_virtual
                        # runtime情報がない古い行では、現在投手列にTeam Virtualを補完する
                        current_pitcher = current_pitcher_by_seq.get(seq, "")
                        if current_pitcher and current_pitcher in pv_map and not pv_map.get(current_pitcher):
                            pv_map[current_pitcher] = virtual
                        score_cells, score_cursor = self._debug_trace_score_cells(ap, team_score_judgments, pitcher_score_judgments, score_cursor, pitcher_by_runner)
                        row = [title, seq, self._short_event(getattr(src, "raw_text", "")), actual, virtual] + [pv_map.get(name, "") for name in pitcher_columns] + [self._delta(ap, vp)] + score_cells
                        last_actual = actual or last_actual
                        last_virtual = virtual or last_virtual

                    last_pitcher_virtuals.update({k: v for k, v in pv_map.items() if v})
                    ws.append(row)
                    rr = ws.max_row
                    for c in ws[rr]:
                        c.alignment = Alignment(vertical="top", wrap_text=True)
                        c.border = self.border


    def _collect_pitcher_columns(self, day):
        """DebugTraceで投手別Virtual列に使う投手名を、試合全体から収集する。"""
        columns = []
        seen = set()
        for game in getattr(day, "games", []):
            for half in getattr(game.analysis, "halves", []):
                display_map = self._pitcher_display_map_for_half(half)
                for name in display_map.values():
                    if name and name not in seen:
                        seen.add(name)
                        columns.append(name)
                for report in [
                    getattr(half, "actual_report", None),
                    getattr(half, "virtual_report", None),
                    getattr(half, "pitcher_virtual_report", None),
                ]:
                    for row in getattr(report, "pitcher_runtime_debug", []) or []:
                        parts = str(row).split(",", 8)
                        if len(parts) == 9:
                            for pid in self._pitcher_ids_from_all_virtuals(parts[7]):
                                name = self._display_pitcher(pid, display_map)
                                if name and name not in seen:
                                    seen.add(name)
                                    columns.append(name)
                            name = self._display_pitcher(parts[2], display_map)
                            if name and name not in seen:
                                seen.add(name)
                                columns.append(name)
        # 実投手名が1人でも判明している場合は、表示専用のプレースホルダー P は列に出さない。
        real_columns = [c for c in columns if str(c).strip() and str(c).strip() != "P"]
        return real_columns or columns or ["P"]

    def _clean_pitcher_name(self, name: str) -> str:
        import re
        s = str(name or "").strip().lstrip("*").strip()
        # TextLiveの投手交代文から混入しやすい記号・投球数を除去
        s = s.lstrip("】］]）) ")
        s = re.sub(r"[（(][^）)]*球[^）)]*[）)]", "", s)
        s = re.sub(r"[（(][^）)]*[）)]", "", s)
        s = s.strip(" 、。:：->→－-")
        return s or str(name or "").strip().lstrip("*")

    def _pitcher_display_map_for_half(self, half):
        """内部ID(P等)を表示名へ変換する。

        初期投手は入力時点では P で始まるため、最初の投手交代ログの
        old_pitcher を P の表示名として採用する。
        """
        mapping = {}
        first_old = ""
        for row in getattr(half.actual_report, "pitcher_changes", []) or []:
            parts = str(row).split(",", 3)
            if len(parts) >= 3:
                old = parts[1].strip()
                new = parts[2].strip()
                old_name = self._clean_pitcher_name(old)
                new_name = self._clean_pitcher_name(new)
                if old and old != "P" and old_name and not first_old:
                    first_old = old_name
                if old:
                    mapping[old] = old_name
                if new:
                    mapping[new] = new_name
        if first_old:
            mapping["P"] = first_old
        else:
            mapping.setdefault("P", "P")

        # Runtime側にしか出ない投手IDも表示対象に含める
        for report in [
            getattr(half, "actual_report", None),
            getattr(half, "virtual_report", None),
            getattr(half, "pitcher_virtual_report", None),
        ]:
            for row in getattr(report, "pitcher_runtime_debug", []) or []:
                parts = str(row).split(",", 8)
                if len(parts) == 9:
                    cur = parts[2].strip().lstrip("*")
                    if cur:
                        mapping.setdefault(cur, mapping.get(cur, self._clean_pitcher_name(cur)))
                    for pid in self._pitcher_ids_from_all_virtuals(parts[7]):
                        mapping.setdefault(pid, mapping.get(pid, self._clean_pitcher_name(pid)))
        return mapping

    def _display_pitcher(self, pitcher_id: str, mapping: dict[str, str]) -> str:
        key = str(pitcher_id or "").strip().lstrip("*")
        return mapping.get(key, self._clean_pitcher_name(key))

    def _pitcher_ids_from_all_virtuals(self, text: str) -> list[str]:
        ids = []
        for part in str(text or "").split("|"):
            if "=" not in part:
                continue
            pid = part.split("=", 1)[0].strip().lstrip("*")
            if pid:
                ids.append(pid)
        return ids

    def _parse_all_pitcher_virtuals(self, text: str, mapping: dict[str, str]) -> dict[str, str]:
        out = {}
        for part in str(text or "").split("|"):
            if "=" not in part:
                continue
            raw_id, state = part.split("=", 1)
            name = self._display_pitcher(raw_id, mapping)
            if name:
                out[name] = state.strip()
        return out

    def _display_all_pitcher_virtuals(self, text: str, mapping: dict[str, str]) -> str:
        """RuntimeDebugの All Pitcher Virtuals も内部IDではなく実名で表示する。"""
        parts = []
        for part in str(text or "").split("|"):
            if "=" not in part:
                cleaned = part.strip()
                if cleaned:
                    parts.append(cleaned)
                continue
            raw_id, state = part.split("=", 1)
            active = "*" if raw_id.strip().startswith("*") else ""
            name = self._display_pitcher(raw_id, mapping)
            parts.append(f"{active}{name}={state.strip()}")
        return " | ".join(parts)

    # ------------------------------------------------------------------
    # Phoenix V3.0 Sprint03: RunDetail
    # ------------------------------------------------------------------
    def _write_run_detail(self, ws, day):
        """得点1点ごとの責任明細を出力する。

        Sprint04-2では、採用判定はTeam Virtualのまま固定しつつ、
        Team/Pitcher Virtual判定を横並びにして VirtualDiff へ渡す。
        """
        headers = [
            "GameNo", "Game", "Inning", "Run No", "Scored Runner ID", "Scored Runner",
            "Charged Pitcher", "Run Charged", "Earned Run", "Judgment", "Reason",
            "Team Judgment", "Team Reason", "Pitcher Judgment", "Pitcher Reason",
            "Adopted Source", "Diff", "Virtual Source", "RunnerText"
        ]
        self._header(ws, headers)
        run_rows = []
        for game in getattr(day, "games", []):
            for half in getattr(game.analysis, "halves", []):
                title = str(getattr(half, "title", "") or "").replace("Actual ", "").replace("Virtual ", "")
                display_map = self._pitcher_display_map_for_half(half)
                pitcher_by_runner = self._runner_pitcher_map(getattr(half.actual_report, "runner_history", []), display_map)
                name_by_runner = self._runner_name_map(getattr(half.actual_report, "runner_history", []))

                compare = getattr(half, "compare_result", None)
                adopted_judgments = list(getattr(compare, "judgments", []) or [])
                team_judgments = list(getattr(compare, "team_judgments", []) or adopted_judgments)
                pitcher_judgments = list(getattr(compare, "pitcher_judgments", []) or adopted_judgments)
                team_by_no = {getattr(j, "score_no", i + 1): j for i, j in enumerate(team_judgments)}
                pitcher_by_no = {getattr(j, "score_no", i + 1): j for i, j in enumerate(pitcher_judgments)}
                adopted_source = getattr(compare, "adopted_source", getattr(compare, "virtual_source", "Team Virtual"))

                for j in adopted_judgments:
                    score_no = getattr(j, "score_no", "")
                    tj = team_by_no.get(score_no, j)
                    pj = pitcher_by_no.get(score_no, j)
                    runner_text = str(getattr(j, "runner_text", "") or "")
                    runner_id = self._runner_id_from_text(runner_text)
                    runner_name = self._runner_name_from_text(runner_text) or name_by_runner.get(runner_id, "")
                    charged_pitcher = pitcher_by_runner.get(runner_id, "")
                    judgment = str(getattr(j, "judgment", "") or "")
                    earned = 1 if judgment == "自責点" else 0
                    team_judgment = str(getattr(tj, "judgment", "") or "")
                    pitcher_judgment = str(getattr(pj, "judgment", "") or "")
                    team_reason = str(getattr(tj, "reason", "") or "")
                    pitcher_reason = str(getattr(pj, "reason", "") or "")
                    diff = "DIFF" if (team_judgment != pitcher_judgment or team_reason != pitcher_reason) else ""
                    row = [
                        getattr(game, "game_no", ""), getattr(game, "game_name", ""), title,
                        score_no, runner_id, runner_name, charged_pitcher,
                        1, earned, judgment, getattr(j, "reason", ""),
                        team_judgment, team_reason, pitcher_judgment, pitcher_reason,
                        adopted_source, diff, getattr(compare, "virtual_source", "Team Virtual"), runner_text,
                    ]
                    ws.append(row)
                    run_rows.append({
                        "game_no": getattr(game, "game_no", ""),
                        "game": getattr(game, "game_name", ""),
                        "inning": title,
                        "run_no": score_no,
                        "runner_id": runner_id,
                        "runner_name": runner_name,
                        "pitcher": charged_pitcher or "(不明)",
                        "run": 1,
                        "earned": earned,
                        "judgment": judgment,
                        "reason": getattr(j, "reason", ""),
                        "team_judgment": team_judgment,
                        "team_reason": team_reason,
                        "pitcher_judgment": pitcher_judgment,
                        "pitcher_reason": pitcher_reason,
                        "adopted_source": adopted_source,
                        "diff": diff,
                        "virtual_source": getattr(compare, "virtual_source", "Team Virtual"),
                        "runner_text": runner_text,
                    })
        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = self.border
        # 自責点/非自責点を軽く色分け
        for r in range(2, ws.max_row + 1):
            judgment = str(ws.cell(r, 10).value or "")
            if judgment == "自責点":
                ws.cell(r, 10).fill = self.ok_fill
                ws.cell(r, 9).fill = self.ok_fill
            elif "非自責" in judgment:
                ws.cell(r, 10).fill = self.warn_fill
                ws.cell(r, 9).fill = self.warn_fill
            if str(ws.cell(r, 17).value or "") == "DIFF":
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = self.ng_fill
        return run_rows

    def _write_virtual_diff(self, ws, run_rows):
        """Team Virtual判定とPitcher Virtual判定の差分だけを表示する開発用シート。"""
        self._header(ws, [
            "GameNo", "Game", "Inning", "Run No", "Scored Runner", "Charged Pitcher",
            "Team Judgment", "Team Reason", "Pitcher Judgment", "Pitcher Reason", "Adopted Source", "Status"
        ])
        diff_rows = [r for r in run_rows if r.get("diff")]
        if not diff_rows:
            ws.append(["", "", "", "", "", "", "", "", "", "", "Team Virtual", "差分なし"])
        else:
            for r in diff_rows:
                ws.append([
                    r.get("game_no", ""), r.get("game", ""), r.get("inning", ""), r.get("run_no", ""),
                    r.get("runner_name", "") or r.get("runner_id", ""), r.get("pitcher", ""),
                    r.get("team_judgment", ""), r.get("team_reason", ""),
                    r.get("pitcher_judgment", ""), r.get("pitcher_reason", ""),
                    r.get("adopted_source", ""), "要確認",
                ])
        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = self.border
        for r in range(2, ws.max_row + 1):
            status = str(ws.cell(r, 12).value or "")
            fill = self.ok_fill if status == "差分なし" else self.ng_fill
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill

    def _write_team_pitcher_summary(self, ws, run_rows):
        """RunDetailを集計して、Teamと投手別の失点・自責点を表示する。"""
        self._header(ws, ["Type", "Name", "Runs Charged", "Earned Runs", "Unearned Runs", "Source"])
        total_runs = sum(int(r.get("run", 0) or 0) for r in run_rows)
        total_earned = sum(int(r.get("earned", 0) or 0) for r in run_rows)
        ws.append(["Team", "Team", total_runs, total_earned, total_runs - total_earned, "RunDetail集計"])

        by_pitcher = {}
        for r in run_rows:
            p = r.get("pitcher") or "(不明)"
            by_pitcher.setdefault(p, {"runs": 0, "earned": 0})
            by_pitcher[p]["runs"] += int(r.get("run", 0) or 0)
            by_pitcher[p]["earned"] += 1 if r.get("pitcher_judgment") == "自責点" else 0
        for pitcher, v in sorted(by_pitcher.items()):
            ws.append(["Pitcher", pitcher, v["runs"], v["earned"], v["runs"] - v["earned"], "RunDetail集計"])

        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = self.border
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == "Team":
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = self.first_fill

    def _runner_pitcher_map(self, runner_history, display_map=None):
        out = {}
        display_map = display_map or {}
        for line in runner_history or []:
            parts = str(line).split(",", 7)
            if len(parts) >= 3:
                out[parts[0]] = self._display_pitcher(parts[2], display_map)
        return out

    def _runner_name_map(self, runner_history):
        out = {}
        for line in runner_history or []:
            parts = str(line).split(",", 7)
            if len(parts) >= 2:
                out[parts[0]] = parts[1]
        return out

    def _runner_id_from_text(self, text: str) -> str:
        if not text:
            return ""
        return str(text).split(":", 1)[0].strip()

    def _runner_name_from_text(self, text: str) -> str:
        if not text or ":" not in text:
            return ""
        rest = text.split(":", 1)[1]
        return rest.split("/", 1)[0].strip()

    def _write_pitcher_runtime_debug(self, ws, day):
        self._header(ws, ["GameNo", "Game", "Half", "Mode", "Seq", "Event", "Current Pitcher", "Pitcher Virtual", "Virtual Outs", "Responsible Runner Count", "Inherited Runner Count", "All Pitcher Virtuals", "Note"])
        for game in getattr(day, "games", []):
            for half in getattr(game.analysis, "halves", []):
                title = str(getattr(half, "title", "") or "").replace("Actual ", "").replace("Virtual ", "")
                display_map = self._pitcher_display_map_for_half(half)
                for mode, report in [
                    ("Actual", half.actual_report),
                    ("Virtual", half.virtual_report),
                    ("PitcherVirtual", getattr(half, "pitcher_virtual_report", None)),
                ]:
                    for row in getattr(report, "pitcher_runtime_debug", []):
                        parts = row.split(",", 8)
                        if len(parts) == 9:
                            parts = list(parts)
                            parts[2] = self._display_pitcher(parts[2], display_map)
                            parts[7] = self._display_all_pitcher_virtuals(parts[7], display_map)
                            ws.append([getattr(game, "game_no", ""), getattr(game, "game_name", ""), title, mode] + parts)
                        else:
                            # Sprint01形式との互換用
                            old_parts = row.split(",", 7)
                            if len(old_parts) == 8:
                                old_parts = list(old_parts)
                                old_parts[2] = self._display_pitcher(old_parts[2], display_map)
                                ws.append([getattr(game, "game_no", ""), getattr(game, "game_name", ""), title, mode] + old_parts[:7] + ["", old_parts[7]])
        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = self.border

    def _short_event(self, line: str) -> str:
        from src.event.event_labeler import short_event_label
        return short_event_label(line)
    def _delta(self, ap, vp) -> str:
        items = []
        if vp is not None:
            try:
                diff = int(getattr(vp, "outs_after", 0) or 0) - int(getattr(vp, "outs_before", 0) or 0)
            except Exception:
                diff = 0
            if diff > 0:
                items.append(f"Virtual Out +{diff}")
            if getattr(vp, "scored_text", []):
                items.append("Runner Score")
        if ap is not None and getattr(ap, "scored_text", []) and "Runner Score" not in items:
            items.append("Runner Score")
        if ap is not None:
            before = str(getattr(ap, "before_text", ""))
            after = str(getattr(ap, "after_text", ""))
            if before != after and "Runner Score" not in items:
                if "一塁[空]" in before and "一塁[空]" not in after:
                    items.append("Runner Add")
                else:
                    items.append("Runner Advance")
        return "\n".join(items)

    def _debug_trace_score_cells(self, ap, team_judgments, pitcher_judgments, score_cursor: int, pitcher_by_runner: dict[str, str]):
        """DebugTrace 1シート検証用の得点判定セルを作る。

        Actualの得点発生行にだけ、Team判定とPitcher判定を並べて表示する。
        Pitcher判定が自責点の場合は責任投手名を表示し、
        Team判定が非自責点でも投手自責点であることを見落とさないようにする。
        """
        if ap is None:
            return ["", "", "", "", ""], score_cursor
        scored_list = list(getattr(ap, "scored_text", []) or [])
        if not scored_list:
            return ["", "", "", "", ""], score_cursor

        score_lines = []
        team_judgment_lines = []
        team_detail_lines = []
        pitcher_judgment_lines = []
        pitcher_detail_lines = []
        for scored in scored_list:
            tj = team_judgments[score_cursor] if score_cursor < len(team_judgments) else None
            pj = pitcher_judgments[score_cursor] if score_cursor < len(pitcher_judgments) else tj
            score_cursor += 1
            score_no = getattr(tj or pj, "score_no", score_cursor) if (tj or pj) else score_cursor
            runner_text = str(getattr(tj or pj, "runner_text", "") or scored or "")
            runner_id = self._runner_id_from_text(runner_text)
            pitcher = pitcher_by_runner.get(runner_id, "")
            team_judgment = str(getattr(tj, "judgment", "") or "") if tj else ""
            team_reason = str(getattr(tj, "reason", "") or "") if tj else ""
            pitcher_judgment = str(getattr(pj, "judgment", "") or "") if pj else ""
            pitcher_reason = str(getattr(pj, "reason", "") or "") if pj else ""

            score_lines.append(f"{score_no}点目")
            team_judgment_lines.append(team_judgment)
            team_detail_lines.append(self._score_detail_for_judgment(team_judgment, team_reason, runner_text, pitcher))
            pitcher_judgment_lines.append(pitcher_judgment)
            pitcher_detail_lines.append(self._score_detail_for_judgment(pitcher_judgment, pitcher_reason, runner_text, pitcher))
        return [
            "\n".join(score_lines),
            "\n".join(team_judgment_lines),
            "\n".join(team_detail_lines),
            "\n".join(pitcher_judgment_lines),
            "\n".join(pitcher_detail_lines),
        ], score_cursor

    def _score_detail_for_judgment(self, judgment: str, reason: str, runner_text: str, pitcher: str) -> str:
        if judgment == "自責点":
            return pitcher or "責任投手不明"
        return self._short_score_reason(reason, runner_text)

    def _short_score_reason(self, reason: str, runner_text: str = "") -> str:
        text = f"{reason} {runner_text}"
        if "Virtual進塁" in text and ("生還不能" in text or "不能" in text):
            return "Virtual生還不能"
        if "Virtual進塁" in text and "判定不能" in text:
            return "要確認"
        if "Virtual" in text or "3アウト" in text:
            return "Virtual3アウト"
        if "reached=tiebreak" in text or "タイブレーク" in text:
            return "タイブレーク走者"
        if "score_cause=field_error" in text or "失策により生還" in text or "悪送球生還" in text:
            return "失策生還"
        if "score_cause=passed_ball" in text or "捕逸により生還" in text or "捕逸生還" in text:
            return "捕逸生還"
        if "field_error" in text or "失策" in text or "落球" in text or "悪送球" in text or "ファンブル" in text:
            return "失策出塁"
        if "fielder_choice" in text or "野選" in text:
            return "野選"
        if "passed_ball" in text or "捕逸" in text:
            return "捕逸"
        if "wild_pitch" in text or "暴投" in text:
            return "暴投"
        if "継承" in text:
            return "継承走者"
        if "自責対象外" in text:
            return "自責対象外出塁"
        return str(reason or "")[:24]

    def _write_summary(self, ws, day):
        from src.version import VERSION
        from src.config import USE_PITCHER_VIRTUAL
        self._header(ws, ["項目", "値"])
        rows = [
            ["Version", VERSION],
            ["Use Pitcher Virtual", str(USE_PITCHER_VIRTUAL)],
            ["目的", "RunDetailを中心に、何回何点目が誰の失点・自責点かを見える化"],
            ["VirtualDiff", "Team Virtual判定とPitcher Virtual判定の差分確認（採用判定はTeam固定）"],
            ["総試合数", getattr(day, "total_games", "")],
            ["総得点数", getattr(day, "total_scores", "")],
            ["補正対象", getattr(day, "total_work_items", "")],
            ["使い方", "PlayTraceで★FIRST_DIFFの行を見る。そこが最初にVirtualがずれた候補です。"],
        ]
        for r in rows:
            ws.append(r)
        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = self.border

    def _write_play_trace(self, ws, day):
        headers = [
            "GameNo", "Game", "Half", "Seq", "RawText",
            "ActualBefore", "ActualAfter", "ActualOutBefore", "ActualOutAfter", "ActualMoves", "ActualScores", "ActualWarnings",
            "VirtualBefore", "VirtualAfter", "VirtualOutBefore", "VirtualOutAfter", "VirtualMoves", "VirtualScores", "VirtualWarnings", "VirtualNotes",
            "OutDiff", "BaseDiff", "Status",
        ]
        self._header(ws, headers)

        first_rows = []
        for game in getattr(day, "games", []):
            for half in getattr(game.analysis, "halves", []):
                actual_by_seq = {getattr(p, "seq", None): p for p in getattr(half.actual_report, "plays", [])}
                virtual_by_seq = {getattr(p, "seq", None): p for p in getattr(half.virtual_report, "plays", [])}
                seqs = sorted({s for s in set(actual_by_seq) | set(virtual_by_seq) if s is not None})

                found_first = False
                for seq in seqs:
                    ap = actual_by_seq.get(seq)
                    vp = virtual_by_seq.get(seq)
                    if not ap and not vp:
                        continue
                    raw = getattr(ap or vp, "raw_text", "")
                    a_after = getattr(ap, "after_text", "") if ap else ""
                    v_after = getattr(vp, "after_text", "") if vp else ""
                    a_out = getattr(ap, "outs_after", "") if ap else ""
                    v_out = getattr(vp, "outs_after", "") if vp else ""
                    out_diff = self._num(v_out) - self._num(a_out) if self._is_num(v_out) and self._is_num(a_out) else ""
                    base_diff = "OK" if self._norm_base(a_after) == self._norm_base(v_after) else "DIFF"
                    status = "OK"
                    if out_diff not in (0, "") or base_diff == "DIFF":
                        status = "DIFF"
                    if not found_first and status == "DIFF":
                        status = "★FIRST_DIFF"
                        found_first = True
                        first_rows.append([
                            getattr(game, "game_no", ""), getattr(game, "game_name", ""), getattr(half, "title", ""), seq,
                            raw, getattr(ap, "before_text", "") if ap else "", a_after, a_out,
                            getattr(vp, "before_text", "") if vp else "", v_after, v_out,
                            out_diff, base_diff,
                        ])

                    row = [
                        getattr(game, "game_no", ""), getattr(game, "game_name", ""), getattr(half, "title", ""), seq, raw,
                        getattr(ap, "before_text", "") if ap else "", a_after,
                        getattr(ap, "outs_before", "") if ap else "", a_out,
                        self._join(getattr(ap, "moves_text", [])) if ap else "",
                        self._join(getattr(ap, "scored_text", [])) if ap else "",
                        self._join(getattr(ap, "warnings", [])) if ap else "",
                        getattr(vp, "before_text", "") if vp else "", v_after,
                        getattr(vp, "outs_before", "") if vp else "", v_out,
                        self._join(getattr(vp, "moves_text", [])) if vp else "",
                        self._join(getattr(vp, "scored_text", [])) if vp else "",
                        self._join(getattr(vp, "warnings", [])) if vp else "",
                        self._join(getattr(vp, "notes", [])) if vp else "",
                        out_diff, base_diff, status,
                    ]
                    ws.append(row)
                    r = ws.max_row
                    for c in ws[r]:
                        c.alignment = Alignment(vertical="top", wrap_text=True)
                        c.border = self.border
                    if status == "OK":
                        ws.cell(r, 23).fill = self.ok_fill
                    elif status == "★FIRST_DIFF":
                        for col in range(1, 24):
                            ws.cell(r, col).fill = self.first_fill
                    else:
                        ws.cell(r, 23).fill = self.ng_fill
        return first_rows

    def _write_first_difference(self, ws, rows):
        self._header(ws, [
            "GameNo", "Game", "Half", "Seq", "RawText",
            "ActualBefore", "ActualAfter", "ActualOutAfter",
            "VirtualBefore", "VirtualAfter", "VirtualOutAfter",
            "OutDiff", "BaseDiff", "見方",
        ])
        if not rows:
            ws.append(["", "", "", "", "差異は検出されませんでした", "", "", "", "", "", "", "", "", ""])
        else:
            for r in rows:
                ws.append(r + ["この行が、その半イニングで最初にActualとVirtualがずれた候補です。"])
        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = self.border
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = self.first_fill

    def _write_score_judgments(self, ws, day):
        self._header(ws, ["GameNo", "Game", "Half", "ScoreNo", "Judgment", "Reason", "RunnerText", "Confidence"])
        for game in getattr(day, "games", []):
            for half in getattr(game.analysis, "halves", []):
                for j in getattr(half.compare_result, "judgments", []):
                    ws.append([
                        getattr(game, "game_no", ""), getattr(game, "game_name", ""), getattr(half, "title", ""),
                        getattr(j, "score_no", ""), getattr(j, "judgment", ""), getattr(j, "reason", ""),
                        getattr(j, "runner_text", ""), getattr(j, "confidence", ""),
                    ])
        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = self.border

    def _write_warnings(self, ws, day):
        self._header(ws, ["GameNo", "Game", "Half", "Mode", "Seq", "RawText", "Warning"])
        for game in getattr(day, "games", []):
            for half in getattr(game.analysis, "halves", []):
                for mode, report in [
                    ("Actual", half.actual_report),
                    ("Virtual", half.virtual_report),
                    ("PitcherVirtual", getattr(half, "pitcher_virtual_report", None)),
                ]:
                    for p in getattr(report, "plays", []):
                        for w in getattr(p, "warnings", []):
                            ws.append([
                                getattr(game, "game_no", ""), getattr(game, "game_name", ""), getattr(half, "title", ""),
                                mode, getattr(p, "seq", ""), getattr(p, "raw_text", ""), w,
                            ])
        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = self.border
        for r in range(2, ws.max_row + 1):
            ws.cell(r, 7).fill = self.warn_fill

    def _autosize(self, ws):
        caps = {1: 10, 2: 28, 3: 12, 4: 8, 5: 52}
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 8
            for cell in ws[letter]:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(80, max(len(x) for x in v.splitlines() or [""])))
            ws.column_dimensions[letter].width = min(caps.get(col, 36), max_len + 2)
        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 22 if r == 1 else 38

    def _join(self, value):
        if value is None:
            return ""
        if isinstance(value, (list, tuple)):
            return "\n".join(self._normalize_note_label(str(x)) for x in value)
        return self._normalize_note_label(str(value))

    def _normalize_note_label(self, text: str) -> str:
        """
        V3.1 Quality06:
        Virtual上に存在しないActual走者の後続移動は、判定上は自然差分。
        Runner側で V自然除外 にしていても、古いNoteや途中生成物が残ると
        DebugReport上に V安全除外 と表示され、警告に見える。
        DebugReport出力時にも最終正規化して、表示品質を安定させる。
        """
        if text.startswith("V安全除外:") and "Virtual上に走者なし" in text:
            return text.replace("V安全除外:", "V自然除外:", 1)
        return text

    def _norm_base(self, text):
        text = str(text or "")
        for p in ["走者:", "Bases:", "base:"]:
            if p in text:
                text = text.split(p, 1)[-1]
        return text.replace(" ", "").replace("　", "")

    def _is_num(self, v):
        try:
            int(v)
            return True
        except Exception:
            return False

    def _num(self, v):
        try:
            return int(v)
        except Exception:
            return 0
