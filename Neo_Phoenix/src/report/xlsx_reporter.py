from __future__ import annotations

from pathlib import Path
import csv


class XlsxReporter:
    """
    Stable Sprint2-2b

    openpyxlを使って、CSVレポートを1冊のExcelへ統合する。
    自前XML生成は廃止。
    """

    SHEET_FILES = [
        ("Summary", None),
        ("ScoreJudgments", "score_judgments.csv"),
        ("Reviews", "reviews.csv"),
        ("PitcherSummary", "pitcher_summary.csv"),
        ("DebugTrace", "debug_trace.csv"),
        ("PitcherRuntimeDebug", "pitcher_runtime_debug.csv"),
        ("RunnerHistory", "runner_history.csv"),
        ("PlayDebug", "play_debug.csv"),
        ("DailyCheck", None),
    ]

    WIDTHS = {
        "Summary": [22, 32],
        "ScoreJudgments": [14, 10, 60, 16, 70, 12],
        "Reviews": [14, 14, 20, 70, 70, 12],
        "PitcherSummary": [18, 10, 10, 10, 14],
        "DebugTrace": [10, 6, 16, 14, 16, 18, 28],
        "PitcherRuntimeDebug": [12, 10, 6, 14, 18, 18, 12, 22, 20, 28],
        "RunnerHistory": [14, 14, 18, 18, 12, 12, 10, 80],
        "PlayDebug": [14, 12, 8, 80, 45, 80, 45],
        "DailyCheck": [18, 12, 16, 34, 20, 70],
    }

    def write_from_reports_dir(self, reports_dir: str | Path, xlsx_path: str | Path, summary_rows: list[list[str]] | None = None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError as e:
            raise RuntimeError(
                "openpyxl がインストールされていません。\\n"
                "次のコマンドを実行してください: pip install openpyxl"
            ) from e

        reports_dir = Path(reports_dir)
        xlsx_path = Path(xlsx_path)
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # default sheet is reused for first sheet
        default_ws = wb.active

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        sub_fill = PatternFill("solid", fgColor="E2F0D9")
        red_fill = PatternFill("solid", fgColor="FCE4D6")
        yellow_fill = PatternFill("solid", fgColor="FFF2CC")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for idx, (sheet_name, csv_name) in enumerate(self.SHEET_FILES):
            if idx == 0:
                ws = default_ws
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(sheet_name)

            if sheet_name == "DailyCheck":
                rows = self._build_daily_check_rows(reports_dir, summary_rows)
            elif csv_name is None:
                rows = summary_rows or [["項目", "値"], ["Status", "Generated"]]
            else:
                rows = self._read_csv(reports_dir / csv_name)

            self._write_rows(ws, rows)
            self._style_sheet(ws, sheet_name, header_fill, sub_fill, red_fill, yellow_fill, border, Font, Alignment, get_column_letter)

        wb.save(xlsx_path)

    def _build_daily_check_rows(self, reports_dir: Path, summary_rows: list[list[str]] | None = None) -> list[list[str]]:
        score_rows = self._read_csv(reports_dir / "score_judgments.csv")
        review_rows = self._read_csv(reports_dir / "reviews.csv")
        pitcher_rows = self._read_csv(reports_dir / "pitcher_summary.csv")

        def summary_value(key: str, default: str = "") -> str:
            for row in summary_rows or []:
                if len(row) >= 2 and row[0] == key:
                    return str(row[1])
            return default

        total = summary_value("総得点", "0")
        earned = summary_value("自責点", "0")
        unearned = summary_value("非自責点", "0")
        candidate = summary_value("自責点候補", "0")
        reviews = summary_value("Review件数", "0")
        target = summary_value("解析対象", summary_value("入力", ""))

        # 今日直すところ：非自責点・自責点候補・WARN/ERROR Reviewのみ
        work_items: list[list[str]] = []

        if len(score_rows) > 1:
            for sr in score_rows[1:]:
                # score_judgments: 半イニング, 点目, 走者, 判定, 理由, 信頼度
                half = sr[0] if len(sr) > 0 else ""
                no = sr[1] if len(sr) > 1 else ""
                runner = sr[2] if len(sr) > 2 else ""
                judge = sr[3] if len(sr) > 3 else ""
                reason = sr[4] if len(sr) > 4 else ""
                conf = sr[5] if len(sr) > 5 else ""

                if judge == "非自責点":
                    label = "🔴 非自責"
                    action = "□ EasyScore補正済"
                elif judge == "自責点候補":
                    label = "🟡 要確認"
                    action = "□ 判定確認"
                else:
                    continue

                location = f"{half} {no}点目".strip()
                # 長すぎるrunnerはDailyCheckでは邪魔なので要約
                runner_short = runner.split(" / ", 1)[0] if runner else ""
                work_items.append([location, label, runner_short, conf, reason, action])

        if len(review_rows) > 1:
            for rr in review_rows[1:]:
                # reviews: 半イニング, レベル, 場所, 内容, 提案, 信頼度
                half = rr[0] if len(rr) > 0 else ""
                level = rr[1] if len(rr) > 1 else ""
                loc = rr[2] if len(rr) > 2 else ""
                msg = rr[3] if len(rr) > 3 else ""
                conf = rr[5] if len(rr) > 5 else ""

                if level in {"WARN", "ERROR"}:
                    location = f"{half} {loc}".strip()
                    work_items.append([location, "🟡 要確認", msg, conf, "Review Engineによる確認推奨", "□ 確認済"])

        rows: list[list[str]] = []
        rows.append(["EasyScoreJudge Phoenix Daily Check", "", "", "", "", ""])
        rows.append(["今日のEasyScore補正作業票", "", "", "", "", ""])
        rows.append(["", "", "", "", "", ""])

        rows.append(["入力", target, "", "確認状態", "□ 未確認  □ 補正済  □ 再確認", ""])
        rows.append(["総得点", total, "自責", earned, "非自責", unearned])
        rows.append(["自責候補", candidate, "Review", reviews, "本日の補正", "□ 完了"])
        rows.append(["", "", "", "", "", ""])

        rows.append(["今日直すところ", "判定", "走者/内容", "信頼度", "理由", "作業"])
        if not work_items:
            rows.append(["", "✅ 補正対象なし", "", "", "", ""])
        else:
            for item in work_items:
                rows.append(item)

        rows.append(["", "", "", "", "", ""])
        rows.append(["投手サマリー", "失点", "自責", "非自責", "自責候補", "備考"])
        for pr in pitcher_rows[1:]:
            rows.append((pr + [""] * 6)[:6])

        rows.append(["", "", "", "", "", ""])
        rows.append(["得点詳細", "判定", "走者", "信頼度", "理由", "確認"])
        if len(score_rows) <= 1:
            rows.append(["", "得点なし", "", "", "", ""])
        else:
            for sr in score_rows[1:]:
                half = sr[0] if len(sr) > 0 else ""
                no = sr[1] if len(sr) > 1 else ""
                runner = sr[2] if len(sr) > 2 else ""
                judge = sr[3] if len(sr) > 3 else ""
                reason = sr[4] if len(sr) > 4 else ""
                conf = sr[5] if len(sr) > 5 else ""
                location = f"{half} {no}点目".strip()
                if judge == "非自責点":
                    label = "🔴 非自責"
                elif judge == "自責点候補":
                    label = "🟡 要確認"
                elif judge == "自責点":
                    label = "🟢 自責"
                else:
                    label = judge
                runner_short = runner.split(" / ", 1)[0] if runner else ""
                rows.append([location, label, runner_short, conf, reason, "□確認済"])

        rows.append(["", "", "", "", "", ""])
        rows.append(["Review詳細", "レベル", "場所", "内容", "信頼度", "確認"])
        if len(review_rows) <= 1:
            rows.append(["", "なし", "", "", "", ""])
        else:
            for rr in review_rows[1:]:
                half = rr[0] if len(rr) > 0 else ""
                level = rr[1] if len(rr) > 1 else ""
                loc = rr[2] if len(rr) > 2 else ""
                msg = rr[3] if len(rr) > 3 else ""
                conf = rr[5] if len(rr) > 5 else ""
                rows.append([half, level, loc, msg, conf, "□確認済"])

        return rows

    def _read_csv(self, path: Path) -> list[list[str]]:
        if not path.exists():
            return [["Message"], [f"Missing file: {path.name}"]]
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return [[str(cell) for cell in row] for row in csv.reader(f)]

    def _write_rows(self, ws, rows: list[list[str]]):
        if not rows:
            ws.append(["No data"])
            return
        for row in rows:
            ws.append(row)

    def _style_daily_check(self, ws, header_fill, sub_fill, red_fill, yellow_fill, border, Font, Alignment):
        # Excel標準風・作業票形式。色は補助、文字は黒中心。
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        dark_header = PatternFill("solid", fgColor="1F4E79")
        light_gray = PatternFill("solid", fgColor="F2F2F2")
        light_red = PatternFill("solid", fgColor="FCE4D6")
        light_yellow = PatternFill("solid", fgColor="FFF2CC")
        light_green = PatternFill("solid", fgColor="E2F0D9")
        light_blue = PatternFill("solid", fgColor="DDEBF7")

        ws.sheet_view.showGridLines = True
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.35
        ws.page_margins.bottom = 0.35

        ws.merge_cells("A1:F1")
        ws.merge_cells("A2:F2")
        ws["A1"].font = Font(bold=True, size=16, color="1F1F1F")
        ws["A2"].font = Font(bold=True, size=11, color="666666")
        ws["A1"].alignment = Alignment(horizontal="left")
        ws["A2"].alignment = Alignment(horizontal="left")

        section_titles = {"今日直すところ", "投手サマリー", "得点詳細", "Review詳細"}

        for row in ws.iter_rows():
            title = str(row[0].value or "")
            is_section = title in section_titles
            is_top_info = row[0].row in {4, 5, 6}

            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                value = str(cell.value or "")

                if is_section:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = dark_header
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif is_top_info:
                    cell.font = Font(bold=True if cell.column in {1, 3, 5} else False, color="000000")
                    if cell.column in {1, 3, 5}:
                        cell.fill = light_gray
                else:
                    cell.font = Font(color="000000")
                    if "🔴" in value or "非自責" in value:
                        cell.fill = light_red
                    elif "🟡" in value or "要確認" in value or "WARN" in value or "ERROR" in value:
                        cell.fill = light_yellow
                    elif "🟢" in value or "自責" in value or "INFO" in value or "補正対象なし" in value:
                        cell.fill = light_green

        # 「今日直すところ」のデータ部分を強調
        in_work_area = False
        for row in ws.iter_rows():
            first = str(row[0].value or "")
            if first == "今日直すところ":
                in_work_area = True
                continue
            if first == "投手サマリー":
                in_work_area = False
            if in_work_area:
                for cell in row:
                    v = str(cell.value or "")
                    if "🔴" in v or "非自責" in v:
                        cell.fill = light_red
                    elif "🟡" in v or "要確認" in v:
                        cell.fill = light_yellow
                    elif "補正対象なし" in v:
                        cell.fill = light_green

        widths = [20, 14, 28, 12, 56, 24]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 18
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20

        # 作業行・詳細行は高め
        for r in range(8, ws.max_row + 1):
            ws.row_dimensions[r].height = 30

        ws.freeze_panes = "A8"
        ws.print_title_rows = "1:8"

    def _style_summary_dashboard(self, ws, header_fill, sub_fill, red_fill, yellow_fill, border, Font, Alignment):
        # Summary rows are key-value pairs. Turn them into a simple dashboard.
        ws.freeze_panes = "A6"
        ws.sheet_view.showGridLines = False

        # Title rows
        ws.insert_rows(1, 3)
        ws["A1"] = "EasyScoreJudge Phoenix"
        ws["A2"] = "Stable Dashboard"
        ws.merge_cells("A1:D1")
        ws.merge_cells("A2:D2")

        ws["A1"].font = Font(bold=True, size=18, color="1F4E79")
        ws["A2"].font = Font(bold=True, size=12, color="666666")
        ws["A1"].alignment = Alignment(horizontal="left")
        ws["A2"].alignment = Alignment(horizontal="left")

        # Header now at row 4
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Body
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            row[0].font = Font(bold=True)

            label = str(row[0].value or "")
            value_cell = row[1]
            if label in {"非自責点", "Review件数"}:
                value_cell.fill = red_fill
            elif label in {"自責点候補", "Review率"}:
                value_cell.fill = yellow_fill
            elif label in {"自責点", "総得点", "解析対象"}:
                value_cell.fill = sub_fill

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 22
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 22

    def _style_sheet(self, ws, sheet_name: str, header_fill, sub_fill, red_fill, yellow_fill, border, Font, Alignment, get_column_letter):
        max_row = ws.max_row
        max_col = ws.max_column

        if sheet_name == "Summary":
            self._style_summary_dashboard(ws, header_fill, sub_fill, red_fill, yellow_fill, border, Font, Alignment)
            return

        if sheet_name == "DailyCheck":
            self._style_daily_check(ws, header_fill, sub_fill, red_fill, yellow_fill, border, Font, Alignment)
            return

        # Header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Body
        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

                value = str(cell.value) if cell.value is not None else ""
                if sheet_name == "ScoreJudgments":
                    if "非自責" in value:
                        cell.fill = red_fill
                    elif "自責点候補" in value:
                        cell.fill = yellow_fill
                elif sheet_name == "Reviews":
                    if value == "ERROR":
                        cell.fill = red_fill
                    elif value == "WARN":
                        cell.fill = yellow_fill
                    elif value == "INFO":
                        cell.fill = sub_fill
                elif sheet_name == "RunnerHistory":
                    if value == "False":
                        cell.fill = red_fill
                    elif value == "True":
                        cell.fill = sub_fill

        # Freeze + filter
        ws.freeze_panes = "A2"
        if max_row >= 1 and max_col >= 1:
            ws.auto_filter.ref = ws.dimensions

        # Widths
        widths = self.WIDTHS.get(sheet_name, [])
        for col_idx in range(1, max_col + 1):
            width = widths[col_idx - 1] if col_idx <= len(widths) else 18
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Row heights
        ws.row_dimensions[1].height = 22
        for row_idx in range(2, max_row + 1):
            ws.row_dimensions[row_idx].height = 36 if sheet_name in {"RunnerHistory", "PlayDebug", "Reviews"} else 20
