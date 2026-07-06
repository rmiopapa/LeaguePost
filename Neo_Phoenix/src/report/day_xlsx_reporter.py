from __future__ import annotations

from pathlib import Path
from datetime import datetime

from src.event.score_event_builder import ScoreEventBuilder


class DayXlsxReporter:
    """
    Phoenix V2.2 Sprint3-2 EventBuilder

    1日最大3試合のEasyScore補正作業票。
    DaySummaryを「今日どこを直すか」に特化したレイアウトへ変更。
    """

    def write(self, day_analysis, out_path: str | Path):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError as e:
            raise RuntimeError(
                "openpyxl がインストールされていません。python -m pip install openpyxl を実行してください。"
            ) from e

        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "DaySummary"

        style = self._make_style(Font, PatternFill, Border, Side, Alignment)
        self._write_day_summary(ws, day_analysis, style, get_column_letter)

        for game in day_analysis.games:
            title = self._safe_sheet_name(f"Game{game.game_no}_{game.game_name}")
            gws = wb.create_sheet(title)
            self._write_game_check(gws, game, style, get_column_letter)

        wb.save(out_path)

    def _make_style(self, Font, PatternFill, Border, Side, Alignment):
        thin = Side(style="thin", color="D9D9D9")
        return {
            "Font": Font,
            "Alignment": Alignment,
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
            "dark": PatternFill("solid", fgColor="1F4E79"),
            "gray": PatternFill("solid", fgColor="F2F2F2"),
            "red": PatternFill("solid", fgColor="FCE4D6"),
            "yellow": PatternFill("solid", fgColor="FFF2CC"),
            "green": PatternFill("solid", fgColor="E2F0D9"),
            "blue": PatternFill("solid", fgColor="DDEBF7"),
            "white": PatternFill("solid", fgColor="FFFFFF"),
        }

    # -----------------------------
    # Collectors
    # -----------------------------

    def _judgment_counts(self, analysis):
        counts = {}
        for half in analysis.halves:
            for j in half.compare_result.judgments:
                counts[j.judgment] = counts.get(j.judgment, 0) + 1
        return counts

    def _score_events_for_game(self, game):
        return ScoreEventBuilder().build_for_game(game)

    def _work_items_for_game(self, game):
        """
        補正ToDo用。
        ReviewのActual #/Virtual #は混ぜない。
        原則として、非自責点と自責点候補だけを作業対象にする。
        自責点は「得点詳細」で確認できる。
        """
        items = []
        builder = ScoreEventBuilder()

        for ev in self._score_events_for_game(game):
            label = builder.label(ev.judgment)
            if ev.judgment == "非自責点":
                action = "□"
            elif ev.judgment == "自責点候補":
                action = "□"
            else:
                continue

            items.append({
                "game": game.game_name,
                "location": ev.location,
                "label": label,
                "runner": ev.runner,
                "confidence": ev.confidence,
                "reason": ev.reason,
                "action": action,
            })

        return items

    def _display_runner(self, runner_text: str) -> str:
        if not runner_text:
            return ""
        text = runner_text.split(" / ", 1)[0]
        if ":" in text:
            text = text.split(":", 1)[1]
        return text.strip()

    def _short_reason(self, reason: str, runner_text: str = "") -> str:
        text = f"{reason} {runner_text}"
        if "reached=tiebreak" in text or "タイブレーク" in text:
            return "タイブレーク走者"
        if "field_error" in text or "失策" in text or "落球" in text or "悪送球" in text:
            return "失策出塁"
        if "fielder_choice" in text or "野選" in text:
            return "野選"
        if "passed_ball" in text or "捕逸" in text:
            return "捕逸"
        if "wild_pitch" in text or "暴投" in text:
            return "暴投"
        if "継承" in text:
            return "継承走者"
        if "Virtual進塁" in text and ("生還不能" in text or "不能" in text):
            return "Virtual生還不能"
        if "Virtual進塁" in text and "判定不能" in text:
            return "要確認"
        if "Virtual" in text or "3アウト" in text:
            return "Virtual3アウト"
        if "自責対象外" in text:
            return "自責対象外出塁"
        return reason[:12] if reason else ""

    def _pitcher_summary_rows(self, analysis):
        summary = {}

        for half in analysis.halves:
            runner_pitcher = {}
            for row in getattr(half.actual_report, "runner_history", []):
                parts = row.split(",", 7)
                if len(parts) == 8:
                    runner_id, name, pitcher, cause, eligible, status, base, hist = parts
                    runner_pitcher[runner_id] = pitcher

            for j in half.compare_result.judgments:
                runner_id = j.runner_text.split(":", 1)[0].strip() if ":" in j.runner_text else ""
                pitcher = runner_pitcher.get(runner_id, "不明")
                summary.setdefault(pitcher, {"失点": 0, "自責": 0, "非自責": 0, "候補": 0})
                summary[pitcher]["失点"] += 1
                if j.judgment == "非自責点":
                    summary[pitcher]["非自責"] += 1
                elif j.judgment == "自責点":
                    summary[pitcher]["自責"] += 1
                else:
                    summary[pitcher]["候補"] += 1

        return [[p, v["失点"], v["自責"], v["非自責"], v["候補"]] for p, v in summary.items()]

    def _review_rows(self, analysis):
        rows = []
        for half in analysis.halves:
            for item in half.review_result.items:
                if item.level == "INFO":
                    continue
                loc = item.location
                if loc.startswith("Actual #") or loc.startswith("Virtual #"):
                    loc = half.title
                rows.append([half.title, item.level, loc, item.message, str(item.confidence)])
        return rows

    # -----------------------------
    # Writers
    # -----------------------------

    def _write_day_summary(self, ws, day, style, get_column_letter):
        Font = style["Font"]
        Alignment = style["Alignment"]

        total_games = len(day.games)
        total_items = 0
        total_confirm = 0
        no_fix_games = 0
        for game in day.games:
            items = self._work_items_for_game(game)
            total_items += len(items)
            total_confirm += sum(1 for x in items if x["label"] == "要確認")
            if not items:
                no_fix_games += 1

        rows = []
        rows.append(["EasyScoreJudge Phoenix V2.5.6 RootSafeFix", "", "", "", "", ""])
        rows.append(["1日最大3試合 自責点確認票", "", "", "", "", ""])
        rows.append(["作成日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "本日の補正完了", "□", ""])
        rows.append(["", "", "", "", "", ""])

        rows.append(["今日の結果", "", "", "", "", ""])
        rows.append(["解析試合", f"{total_games}試合", "補正対象", f"{total_items}件", "要確認", f"{total_confirm}件"])
        rows.append(["補正なし", f"{no_fix_games}試合", "", "", "", ""])
        rows.append(["", "", "", "", "", ""])

        rows.append(["本日の補正ToDo", "内容", "理由", "作業", "", ""])

        if not day.games:
            rows.append(["gamesフォルダにtxtがありません", "", "", "", "", ""])
        else:
            for game in day.games:
                items = self._work_items_for_game(game)
                if not items:
                    rows.append([f"□ {game.game_name}", "補正なし", "", "確認不要", "", ""])
                else:
                    for item in items:
                        content = f"{item['location']}　{item['label']}"
                        rows.append([f"□ {game.game_name}", content, item["reason"], item["action"], "", ""])

        for row in rows:
            ws.append(row)

        ws.merge_cells("A1:F1")
        ws.merge_cells("A2:F2")
        ws["A1"].font = Font(bold=True, size=16, color="1F1F1F")
        ws["A2"].font = Font(bold=True, size=11, color="666666")
        ws["A1"].alignment = Alignment(horizontal="left")
        ws["A2"].alignment = Alignment(horizontal="left")

        self._style_common(ws, style, section_titles={"今日の結果", "本日の補正ToDo"})
        self._style_values(ws, style)

        widths = [22, 30, 18, 14, 10, 10]
        self._apply_widths(ws, widths, get_column_letter)
        self._page_setup(ws)
        ws.freeze_panes = "A9"
        ws.print_title_rows = "1:9"

    def _write_game_check(self, ws, game, style, get_column_letter):
        Font = style["Font"]
        Alignment = style["Alignment"]

        counts = self._judgment_counts(game.analysis)
        work_items = self._work_items_for_game(game)
        pitcher_rows = self._pitcher_summary_rows(game.analysis)
        review_rows = self._review_rows(game.analysis)

        rows = []
        rows.append(["EasyScoreJudge Phoenix Daily Check", "", "", "", "", ""])
        rows.append([f"{game.game_name}：EasyScore補正票", "", "", "", "", ""])
        rows.append(["", "", "", "", "", ""])
        rows.append(["試合名", game.game_name, "", "確認状態", "□未確認  □補正済  □再確認", ""])
        rows.append(["総得点", game.analysis.total_scores, "自責", counts.get("自責点", 0), "非自責", counts.get("非自責点", 0)])
        rows.append(["自責候補", counts.get("自責点候補", 0), "Review", len(review_rows), "本日の補正", "□"])
        rows.append(["", "", "", "", "", ""])

        rows.append(["補正作業", "場所", "判定", "理由", "走者/内容", "作業"])
        if not work_items:
            rows.append(["", "補正対象なし", "", "", "", "確認不要"])
        else:
            for idx, item in enumerate(work_items, 1):
                rows.append([
                    f"{idx}",
                    item["location"],
                    item["label"],
                    item["reason"],
                    item["runner"],
                    "□",
                ])

        rows.append(["", "", "", "", "", ""])
        rows.append(["投手サマリー", "項目", "値", "", "", ""])
        if not pitcher_rows:
            rows.append(["", "失点", 0, "", "", ""])
            rows.append(["", "自責", 0, "", "", ""])
            rows.append(["", "非自責", 0, "", "", ""])
        else:
            for pr in pitcher_rows:
                pitcher, runs, earned, unearned, candidate = pr
                rows.append([pitcher, "失点", runs, "", "", ""])
                rows.append(["", "自責", earned, "", "", ""])
                rows.append(["", "非自責", unearned, "", "", ""])
                if candidate:
                    rows.append(["", "自責候補", candidate, "", "", ""])

        rows.append(["", "", "", "", "", ""])
        rows.append(["得点詳細", "場所", "判定", "理由", "走者", "確認"])
        score_events = self._score_events_for_game(game)
        if not score_events:
            rows.append(["", "", "得点なし", "", "", ""])
        else:
            builder = ScoreEventBuilder()
            for ev in score_events:
                rows.append([ev.location, ev.location, builder.label(ev.judgment), ev.reason, ev.runner, "□"])

        rows.append(["", "", "", "", "", ""])
        rows.append(["Review詳細", "レベル", "場所", "内容", "信頼度", "確認"])
        if not review_rows:
            rows.append(["", "なし", "", "", "", ""])
        else:
            for rr in review_rows:
                rows.append(rr + ["□"])

        for row in rows:
            ws.append(row)

        ws.merge_cells("A1:F1")
        ws.merge_cells("A2:F2")
        ws["A1"].font = Font(bold=True, size=16, color="1F1F1F")
        ws["A2"].font = Font(bold=True, size=11, color="666666")
        ws["A1"].alignment = Alignment(horizontal="left")
        ws["A2"].alignment = Alignment(horizontal="left")

        self._style_common(ws, style, section_titles={"補正作業", "投手サマリー", "得点詳細", "Review詳細"})
        self._style_values(ws, style)

        widths = [10, 22, 14, 18, 34, 10]
        self._apply_widths(ws, widths, get_column_letter)
        self._page_setup(ws)
        ws.freeze_panes = "A8"
        ws.print_title_rows = "1:8"

    # -----------------------------
    # Styles
    # -----------------------------

    def _label_for_judgment(self, judgment: str) -> str:
        if judgment == "非自責点":
            return "非自責"
        if judgment == "自責点候補":
            return "要確認"
        if judgment == "自責点":
            return "自責"
        return judgment

    def _style_common(self, ws, style, section_titles: set[str] | None = None):
        Font = style["Font"]
        Alignment = style["Alignment"]
        section_titles = section_titles or set()

        for row in ws.iter_rows():
            first = str(row[0].value or "")
            is_section = first in section_titles
            is_group = first.startswith("【") and first.endswith("】")
            is_top_info = row[0].row in {3, 4, 5, 6}

            for cell in row:
                cell.border = style["border"]
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if is_section:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = style["dark"]
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif is_group:
                    cell.font = Font(bold=True, color="000000")
                    cell.fill = style["gray"]
                elif is_top_info and cell.column in {1, 3, 5}:
                    cell.font = Font(bold=True, color="000000")
                    cell.fill = style["gray"]
                else:
                    cell.font = Font(color="000000")

    def _style_values(self, ws, style):
        for row in ws.iter_rows():
            for cell in row:
                value = str(cell.value or "")
                if value in {"非自責", "要対応"} or "非自責" in value:
                    cell.fill = style["red"]
                elif value in {"要確認"} or "WARN" in value or "ERROR" in value:
                    cell.fill = style["yellow"]
                elif value in {"自責", "補正対象なし", "補正なし"}:
                    cell.fill = style["green"]

    def _apply_widths(self, ws, widths, get_column_letter):
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 22
        for r in range(8, ws.max_row + 1):
            ws.row_dimensions[r].height = 28

    def _page_setup(self, ws):
        ws.sheet_view.showGridLines = True
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.35
        ws.page_margins.bottom = 0.35

    def _safe_sheet_name(self, name: str) -> str:
        bad = '[]:*?/\\'
        for ch in bad:
            name = name.replace(ch, "_")
        return name[:31]
