from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from src.config import GOLDDATA_EXCLUDED_CASES
from src.game.game_text_reader import GameTextReader
from src.move.models import Move
from src.neo.engine import NeoHalfInningEngine
from src.neo.judgment import NeoScoreJudgmentBuilder

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise ModuleNotFoundError("openpyxl がありません。python -m pip install openpyxl を実行してください。") from exc


SCORE_HEADERS = [
    "case",
    "inning",
    "score_no",
    "golddata",
    "neo",
    "direction",
    "runner",
    "seq",
    "same_seq_virtual_scored",
    "virtual_outs_before",
    "virtual_outs_after",
    "reached_cause_type",
    "score_cause_type",
    "reason",
    "raw_text",
]

PLAY_HEADERS = [
    "case",
    "inning",
    "seq",
    "raw_text",
    "actual_outs_before",
    "actual_outs_after",
    "virtual_outs_before",
    "virtual_outs_after",
    "actual_before",
    "actual_after",
    "virtual_before",
    "virtual_after",
    "actual_scored_runner_ids",
    "virtual_scored_runner_ids",
    "same_seq_scored_runner_ids",
    "actual_moves",
    "virtual_moves",
    "warnings",
]


class NeoDebugReportBuilder:
    def __init__(self, root: Path | str = "regression_cases"):
        self.root = Path(root)
        self.reader = GameTextReader()
        self.judgment_builder = NeoScoreJudgmentBuilder()
        self._setup_styles()

    def build(self, out_path: Path | str, case_filter: str = "") -> Path:
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        score_rows: list[dict[str, Any]] = []
        play_rows: list[dict[str, Any]] = []
        warning_rows: list[dict[str, Any]] = []

        cases = self._case_dirs(case_filter)
        for case_dir in cases:
            try:
                case_score_rows, case_play_rows, case_warning_rows = self._run_case(case_dir)
                score_rows.extend(case_score_rows)
                play_rows.extend(case_play_rows)
                warning_rows.extend(case_warning_rows)
            except Exception as exc:
                warning_rows.append({"case": case_dir.name, "inning": "", "seq": "", "warning": f"case failed: {exc}"})

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_scores = wb.create_sheet("ScoreReview")
        ws_trace = wb.create_sheet("PlayTrace")
        ws_warnings = wb.create_sheet("Warnings")

        self._write_summary(ws_summary, cases, score_rows, play_rows, warning_rows)
        self._write_table(ws_scores, SCORE_HEADERS, score_rows)
        self._write_table(ws_trace, PLAY_HEADERS, play_rows)
        self._write_table(ws_warnings, ["case", "inning", "seq", "warning"], warning_rows)
        self._style_score_review(ws_scores)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            self._autosize(ws)

        wb.save(out_path)
        return out_path

    def _case_dirs(self, case_filter: str) -> list[Path]:
        if case_filter:
            names = {part.strip() for part in case_filter.split(",") if part.strip()}
            return [
                self.root / name
                for name in sorted(names)
                if (self.root / name).is_dir() and name not in GOLDDATA_EXCLUDED_CASES
            ]
        return sorted(
            p for p in self.root.glob("RC*")
            if p.is_dir() and p.name not in GOLDDATA_EXCLUDED_CASES
        )

    def _run_case(self, case_dir: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
        sample_path = case_dir / "sample.txt"
        if not sample_path.exists():
            return [], [], [{"case": case_dir.name, "inning": "", "seq": "", "warning": "sample.txt not found"}]

        expected = self._expected_by_slot(case_dir)
        score_rows: list[dict[str, Any]] = []
        play_rows: list[dict[str, Any]] = []
        warning_rows: list[dict[str, Any]] = []

        for block in self.reader.read(sample_path).half_innings:
            neo = NeoHalfInningEngine(f"Neo {block.title}").run(block.lines)
            judgments = list(self.judgment_builder.build(neo).judgments)
            score_index = 0

            for snap in neo.plays:
                actual_scored = [rid for rid in getattr(snap, "actual_scored_runner_ids", []) or [] if rid]
                virtual_scored = [rid for rid in getattr(snap, "virtual_scored_runner_ids", []) or [] if rid]
                same_seq = sorted(set(actual_scored) & set(virtual_scored))
                facts_by_id = {
                    str(fact.get("id", "")): fact
                    for fact in getattr(snap, "actual_scored_runner_facts", []) or []
                }
                play_rows.append(
                    {
                        "case": case_dir.name,
                        "inning": block.title,
                        "seq": snap.seq,
                        "raw_text": snap.raw_text,
                        "actual_outs_before": snap.actual_outs_before,
                        "actual_outs_after": snap.actual_outs_after,
                        "virtual_outs_before": snap.virtual_outs_before,
                        "virtual_outs_after": snap.virtual_outs_after,
                        "actual_before": snap.actual_before,
                        "actual_after": snap.actual_after,
                        "virtual_before": snap.virtual_before,
                        "virtual_after": snap.virtual_after,
                        "actual_scored_runner_ids": ",".join(actual_scored),
                        "virtual_scored_runner_ids": ",".join(virtual_scored),
                        "same_seq_scored_runner_ids": ",".join(same_seq),
                        "actual_moves": self._format_moves(snap.actual_moves),
                        "virtual_moves": self._format_moves(snap.virtual_moves),
                        "warnings": " / ".join(getattr(snap, "warnings", []) or []),
                    }
                )
                for warning in getattr(snap, "warnings", []) or []:
                    warning_rows.append({"case": case_dir.name, "inning": block.title, "seq": snap.seq, "warning": warning})

                for runner_id in actual_scored:
                    if score_index >= len(judgments):
                        continue
                    judgment = judgments[score_index]
                    score_index += 1
                    slot = (block.title, int(judgment.score_no))
                    gold = expected.get(slot, "")
                    neo_judgment = str(judgment.judgment)
                    fact = facts_by_id.get(str(runner_id), {})
                    score_rows.append(
                        {
                            "case": case_dir.name,
                            "inning": block.title,
                            "score_no": int(judgment.score_no),
                            "golddata": gold,
                            "neo": neo_judgment,
                            "direction": self._direction(gold, neo_judgment),
                            "runner": str(judgment.runner_text),
                            "seq": snap.seq,
                            "same_seq_virtual_scored": str(runner_id) in set(virtual_scored),
                            "virtual_outs_before": snap.virtual_outs_before,
                            "virtual_outs_after": snap.virtual_outs_after,
                            "reached_cause_type": str(fact.get("reached_cause_type", "") or ""),
                            "score_cause_type": str(fact.get("score_cause_type", "") or ""),
                            "reason": str(judgment.reason),
                            "raw_text": snap.raw_text,
                        }
                    )

        return score_rows, play_rows, warning_rows

    def _expected_by_slot(self, case_dir: Path) -> dict[tuple[str, int], str]:
        expected_path = case_dir / "expected.json"
        if not expected_path.exists():
            return {}
        data = json.loads(expected_path.read_text(encoding="utf-8"))
        return {
            (str(item.get("inning", "")), int(item.get("score_no", 0) or 0)): str(item.get("judgment", ""))
            for item in data.get("expected", [])
        }

    def _format_moves(self, moves: list[Move]) -> str:
        parts: list[str] = []
        for mv in moves or []:
            parts.append(
                f"{getattr(mv, 'source', '')}->{getattr(mv, 'target', '')}"
                f"({getattr(mv, 'cause_type', '')}:{getattr(mv, 'reason', '')})"
            )
        return " / ".join(parts)

    def _direction(self, gold: str, neo: str) -> str:
        if not gold:
            return "neo_only"
        if gold == neo:
            return "match"
        return f"{gold}_to_{neo}"

    def _setup_styles(self) -> None:
        self.header_fill = PatternFill("solid", fgColor="1F4E78")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.match_fill = PatternFill("solid", fgColor="E2F0D9")
        self.diff_fill = PatternFill("solid", fgColor="FCE4D6")
        self.warn_fill = PatternFill("solid", fgColor="FFF2CC")
        self.info_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="D9E2F3")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_summary(
        self,
        ws,
        cases: list[Path],
        score_rows: list[dict[str, Any]],
        play_rows: list[dict[str, Any]],
        warning_rows: list[dict[str, Any]],
    ) -> None:
        matched = sum(1 for row in score_rows if row.get("direction") == "match")
        different = sum(1 for row in score_rows if row.get("direction") not in {"match", ""})
        same_seq_after_3 = sum(
            1
            for row in score_rows
            if row.get("same_seq_virtual_scored") is True and int(row.get("virtual_outs_after") or 0) >= 3
        )
        rows = [
            ["Metric", "Value"],
            ["Cases", len(cases)],
            ["Score rows", len(score_rows)],
            ["Score matches", matched],
            ["Score differences", different],
            ["Play rows", len(play_rows)],
            ["Warnings", len(warning_rows)],
            ["Same Seq virtual scored and virtual_outs_after >= 3", same_seq_after_3],
            ["Excluded invalid GoldData cases", ", ".join(sorted(GOLDDATA_EXCLUDED_CASES))],
        ]
        for row in rows:
            ws.append(row)
        self._style_header(ws, 1)

    def _write_table(self, ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
        ws.append(headers)
        self._style_header(ws, 1)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = self.border

    def _style_header(self, ws, row_no: int) -> None:
        for cell in ws[row_no]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.border

    def _style_score_review(self, ws) -> None:
        direction_col = SCORE_HEADERS.index("direction") + 1
        same_seq_col = SCORE_HEADERS.index("same_seq_virtual_scored") + 1
        virtual_after_col = SCORE_HEADERS.index("virtual_outs_after") + 1
        for row_no in range(2, ws.max_row + 1):
            direction = str(ws.cell(row_no, direction_col).value or "")
            same_seq = ws.cell(row_no, same_seq_col).value is True
            try:
                virtual_after = int(ws.cell(row_no, virtual_after_col).value or 0)
            except Exception:
                virtual_after = 0
            fill = self.match_fill if direction == "match" else self.diff_fill
            if same_seq and virtual_after >= 3:
                fill = self.warn_fill
            for cell in ws[row_no]:
                cell.fill = fill

    def _autosize(self, ws) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                text = str(cell.value or "")
                max_len = max(max_len, min(len(text), 80))
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))
        for row in ws.iter_rows():
            for cell in row:
                cell.border = self.border


def main() -> int:
    parser = argparse.ArgumentParser(description="Write Neo_Phoenix debug report as xlsx.")
    parser.add_argument("--root", default="regression_cases")
    parser.add_argument("--xlsx", default="reports/neo_debug_report.xlsx")
    parser.add_argument("--case", default="", help="Comma-separated RC ids, e.g. RC001,RC186")
    args = parser.parse_args()

    out = NeoDebugReportBuilder(args.root).build(args.xlsx, case_filter=args.case)
    print(f"Neo debug report written: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
